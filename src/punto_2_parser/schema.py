"""Lee el Excel de documentación y extrae los campos leaf del Body."""

import logging
from pathlib import Path

import openpyxl

log = logging.getLogger(__name__)

# Formatos que indican agrupadores (no son campos de datos reales)
GROUP_FORMATS = {"Grupo", "GROUP", "GRP"}


def load_body_schema(xlsx_path: Path) -> list[dict]:
    """
    Retorna lista de campos leaf del sheet 'Body':
        [{"name": "campo_N", "pos_ini": int, "pos_fin": int}, ...]

    Excluye filas de tipo Grupo/GROUP/GRP (son contenedores, no datos).
    Las posiciones en el Excel son 1-indexed; se devuelven igual (conversión a 0-index en el parser).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    try:
        ws = wb["Body"]
    except KeyError:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet 'Body' no encontrado en '{xlsx_path.name}'. "
            f"Sheets disponibles: {available}"
        )

    fields = []
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip 2-row header
        if len(row) < 8:
            log.debug("Fila con menos de 8 columnas, ignorando: %s", row)
            continue

        elm, _campo, _niv, formato, _tipo, pos_ini, pos_fin, _long = row[:8]

        if elm is None or pos_ini is None or pos_fin is None:
            continue

        if formato in GROUP_FORMATS:
            log.debug("Saltando agrupador: elm=%s, formato=%s", elm, formato)
            continue

        fields.append({"name": f"campo_{elm}", "pos_ini": int(pos_ini), "pos_fin": int(pos_fin)})

    log.info("Schema cargado: %d campos leaf desde '%s'", len(fields), xlsx_path.name)
    return fields
